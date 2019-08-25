import json
import os
from os.path import join, isfile
from openpyxl import load_workbook
import pandas
import xlsxwriter

out_dir = 'out/'
excel_dir = 'excel_result/'
excel_file = excel_dir + 'comparer_statistics.xlsx'
if not os.path.isfile(excel_file):
    workbook = xlsxwriter.Workbook(excel_file)
    worksheet = workbook.add_worksheet()
    worksheet.merge_range('B1:E1', "Yolo")
    worksheet.merge_range('F1:I1', "SSD")
    worksheet.merge_range('J1:M1', "RCNN")
    worksheet.write('A2', 'Video name')
    worksheet.write('B2', 'mAP')
    worksheet.write('C2', 'Recall')
    worksheet.write('D2', 'Precision')
    worksheet.write('E2', 'Time')
    worksheet.write('F2', 'mAP')
    worksheet.write('G2', 'Recall')
    worksheet.write('H2', 'Precision')
    worksheet.write('I2', 'Time')
    worksheet.write('J2', 'mAP')
    worksheet.write('K2', 'Recall')
    worksheet.write('L2', 'Precision')
    worksheet.write('M2', 'Time')
    workbook.close()

dirs = os.listdir(out_dir)
writer = pandas.ExcelWriter(excel_file, engine='openpyxl')
writer.book = load_workbook(excel_file)
writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
# for sub_dir in dirs:
sub_dir = '/bike'
if not sub_dir.startswith('.'):
    sub_dirs = os.listdir(out_dir + sub_dir)
    for video_name in sub_dirs:
        if not video_name.startswith('.'):
            statistics_files_dir = out_dir + sub_dir + '/' + video_name + '/statistics'
            statistics_files = [f for f in os.listdir(statistics_files_dir) if
                                isfile(join(statistics_files_dir, f))]
            data = []
            data.append(video_name)
            for file_name in statistics_files:
                file = open(statistics_files_dir + '/' + file_name)
                json_dict = json.load(file)
                statistics = json_dict['finalStatistics'][0]
                data.append(statistics['mAP'])
                data.append(statistics['recall'])
                data.append(statistics['precision'])
                data.append(statistics['time'])
            writer.book.worksheets[0].append(data)
writer.save()
